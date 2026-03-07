import streamlit as st
import openpyxl
import io

# 1. Judul Halaman Khusus Data
st.set_page_config(page_title="Olah Data & SIPD", layout="wide")

# 2. Sistem Login (Sama seperti sebelumnya)
# --- SISTEM LOGIN (KATA SANDI) ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔒 Akses Terbatas")
    st.write("Aplikasi ini bersifat privat. Silakan masukkan kata sandi untuk melanjutkan.")
    
    # Membungkus input dan tombol ke dalam st.form agar bisa pakai 'Enter'
    with st.form("login_form"):
        password_input = st.text_input("Kata Sandi:", type="password")
        
        # st.button diganti menjadi st.form_submit_button
        tombol_masuk = st.form_submit_button("Masuk")
        
        if tombol_masuk:
            if password_input == st.secrets["APP_PASSWORD"]: 
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Kata sandi salah!")
    st.stop()

# 3. KONTEN APLIKASI UTAMA (Tanpa Tab)
if st.session_state.authenticated:
    st.title("📊 Mamayo Data Center")
    st.write("Pusat Pengolahan Data Excel dan SIPD")
    
    ## ==========================================
# TAB 10: ALAT EXCEL (TAMBAH/HAPUS PETIK)
# ==========================================
    st.write("Alat otomatis untuk memanipulasi tanda petik (') pada data Excel. Sangat berguna untuk data NIK, No. HP, atau Rekening agar tidak berubah menjadi rumus/angka eksponensial (E+).")
    
    # --- PANDUAN ---
    with st.expander("📖 Buka Panduan Penggunaan", expanded=False):
        st.markdown("""
        **Cara Penggunaan:**
        1. Unggah file Excel (`.xlsx`) Anda di bawah.
        2. Pilih mode apakah ingin **Menambah** atau **Menghapus** tanda petik.
        3. Ketik huruf kolom yang berisi data Anda (Misal: `C`, `D`, `AA`).
        4. Klik tombol Proses dan download hasilnya.
        """)
        
    st.markdown("---")
    
    # 1. Upload File Excel
    file_excel = st.file_uploader("📥 Unggah File Excel (.xlsx)", type=["xlsx"], key="excel_upload")
    
    if file_excel:
        col1, col2 = st.columns(2)
        
        with col1:
            # 2. Input Kolom (Menerjemahkan simpledialog)
            kolom_target = st.text_input("🔠 Ketik Huruf Kolom (Contoh: C):", max_chars=3).upper()
            
        with col2:
            # 3. Pilih Mode
            mode_excel = st.radio("⚙️ Pilih Aksi:", ["+ Tambah Petik Tersembunyi", "- Hapus Semua Petik"], horizontal=True)
            
        if st.button("🚀 PROSES FILE EXCEL", type="primary", use_container_width=True):
            if not kolom_target:
                st.error("⚠️ Mohon isi huruf kolom terlebih dahulu!")
            else:
                with st.spinner("Memproses ribuan baris data Excel..."):
                    try:
                        
                        # Baca file excel dari memori (bukan path folder)
                        wb = openpyxl.load_workbook(file_excel)
                        ws = wb.active
                        
                        # LOGIKA TAMBAH PETIK
                        if mode_excel == "+ Tambah Petik Tersembunyi":
                            for row in range(2, ws.max_row + 1):
                                cell = ws[f"{kolom_target}{row}"]
                                if cell.value is not None:
                                    val_str = str(cell.value).strip()
                                    if val_str.startswith("'"):
                                        val_str = val_str[1:]
                                    cell.value = val_str
                                    cell.quotePrefix = True # Jurus Rahasia VBA
                                    
                            nama_file_baru = f"SiapUpload_{file_excel.name}"
                            pesan_sukses = "✅ Mantap! Petik tersembunyi berhasil ditambahkan."
                            
                        # LOGIKA HAPUS PETIK
                        else:
                            for row in range(2, ws.max_row + 1):
                                cell = ws[f"{kolom_target}{row}"]
                                if cell.value is not None:
                                    val_str = str(cell.value).replace("'", "") # Hancurkan petik
                                    cell.value = val_str
                                    cell.quotePrefix = False 
                                    cell.number_format = '@' # Kunci sebagai teks
                                    
                            nama_file_baru = f"TanpaPetik_{file_excel.name}"
                            pesan_sukses = "✅ Mantap! Seluruh petik berhasil dibersihkan."

                        # Simpan hasil olahan ke dalam memori virtual (BytesIO)
                        output_excel = io.BytesIO()
                        wb.save(output_excel)
                        output_excel.seek(0)
                        
                        st.success(pesan_sukses)
                        
                        # Tombol Download
                        st.download_button(
                            label="📥 Download Hasil Excel",
                            data=output_excel,
                            file_name=nama_file_baru,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                        
                    except Exception as e:

                        st.error(f"❌ Terjadi kesalahan saat membaca file: {e}")
