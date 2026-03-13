import io
import re
import time
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from streamlit_option_menu import option_menu
from supabase import Client, create_client

# ==========================================
# 1. PENGATURAN HALAMAN
# ==========================================
st.set_page_config(
    page_title="Olah Data & SIPD",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="expanded",
)

# ==========================================
# 2. HELPER UI / FORMAT / PLOTLY
# ==========================================
COLOR_PRIMARY = "#2563eb"
COLOR_SECONDARY = "#0ea5e9"
COLOR_SUCCESS = "#16a34a"
COLOR_WARNING = "#f59e0b"
COLOR_DANGER = "#ef4444"
COLOR_VIOLET = "#8b5cf6"
COLOR_TEAL = "#14b8a6"
COLOR_ORANGE = "#f97316"

CHART_PALETTE = [
    COLOR_PRIMARY,
    COLOR_SECONDARY,
    "#22c55e",
    COLOR_WARNING,
    COLOR_VIOLET,
    COLOR_DANGER,
    COLOR_TEAL,
    COLOR_ORANGE,
]

def format_rupiah(x):
    try:
        return f"Rp {float(x):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"

def safe_stem(text):
    return re.sub(r"[^a-zA-Z0-9_-]+", "_", str(text)).strip("_").lower()

def shorten_text(text, limit=42):
    text = str(text)
    return text if len(text) <= limit else text[:limit].rstrip() + "…"

def inject_custom_css():
    st.markdown(
        """
        <style>
        html, body, [class*="css"] {
            font-family: Inter, "Segoe UI", Arial, sans-serif;
        }

        [data-testid="stAppViewContainer"]{
            background:
                radial-gradient(circle at top left, rgba(37,99,235,0.10), transparent 26%),
                radial-gradient(circle at top right, rgba(14,165,233,0.10), transparent 22%),
                linear-gradient(180deg, #f8fbff 0%, #f3f7fc 100%);
        }

        [data-testid="stHeader"]{
            background: rgba(255,255,255,0.70);
            backdrop-filter: blur(10px);
            border-bottom: 1px solid rgba(148,163,184,0.16);
        }

        [data-testid="stSidebar"]{
            background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
            border-right: 1px solid rgba(255,255,255,0.05);
        }

        [data-testid="stSidebar"] *{
            color: #e5eefc;
        }

        .block-container{
            max-width: 1450px;
            padding-top: 1.15rem;
            padding-bottom: 2rem;
        }

        div[data-testid="stMetric"]{
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(148,163,184,0.16);
            border-radius: 22px;
            padding: 16px 18px;
            box-shadow: 0 10px 28px rgba(15,23,42,0.06);
        }

        div[data-testid="stMetricLabel"] p{
            font-weight: 600;
        }

        div[data-testid="stPlotlyChart"]{
            background: rgba(255,255,255,0.95);
            border: 1px solid rgba(148,163,184,0.16);
            border-radius: 24px;
            padding: 10px 12px 8px 12px;
            box-shadow: 0 12px 34px rgba(15,23,42,0.06);
            transition: transform .18s ease, box-shadow .18s ease;
        }

        div[data-testid="stPlotlyChart"]:hover{
            transform: translateY(-2px);
            box-shadow: 0 18px 42px rgba(15,23,42,0.10);
        }

        div[data-baseweb="tab-list"]{
            gap: .55rem;
            margin-bottom: .45rem;
            flex-wrap: wrap;
        }

        div[data-baseweb="tab-list"] button{
            background: rgba(255,255,255,0.84);
            border: 1px solid rgba(148,163,184,0.16);
            border-radius: 999px;
            padding: .35rem .95rem;
        }

        div[data-baseweb="tab-list"] button[aria-selected="true"]{
            background: linear-gradient(135deg, #1d4ed8, #0ea5e9);
            color: white;
            border-color: transparent;
        }

        .hero-box{
            background: linear-gradient(135deg, rgba(29,78,216,0.96), rgba(14,165,233,0.88));
            color: white;
            padding: 1.18rem 1.3rem;
            border-radius: 26px;
            box-shadow: 0 18px 40px rgba(29,78,216,0.22);
            margin-bottom: .95rem;
        }

        .hero-box h3{
            margin: 0;
            font-size: 1.35rem;
            font-weight: 700;
            color: white;
        }

        .hero-box p{
            margin: .42rem 0 0 0;
            color: rgba(255,255,255,0.93);
            font-size: .95rem;
        }

        .glass-box{
            background: rgba(255,255,255,0.88);
            border: 1px solid rgba(148,163,184,0.16);
            border-radius: 22px;
            padding: 1rem 1.05rem;
            box-shadow: 0 10px 30px rgba(15,23,42,0.05);
            margin-bottom: .85rem;
        }

        .section-chip{
            display: inline-block;
            background: rgba(37,99,235,0.10);
            color: #1d4ed8;
            padding: .35rem .75rem;
            border-radius: 999px;
            font-size: .82rem;
            font-weight: 700;
            margin-bottom: .55rem;
        }

        .stButton > button, .stDownloadButton > button {
            border-radius: 999px;
            font-weight: 600;
        }

        .stTextInput > div > div, .stSelectbox > div > div, .stMultiSelect > div > div {
            border-radius: 16px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def render_page_hero(title, subtitle):
    st.markdown(
        f"""
        <div class="hero-box">
            <h3>{title}</h3>
            <p>{subtitle}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

def get_plotly_download_config(filename):
    return {
        "displaylogo": False,
        "responsive": True,
        "toImageButtonOptions": {
            "format": "png",
            "filename": filename,
            "width": None,
            "height": None,
            "scale": 3,
        },
    }

def style_fig(fig, height=420, horizontal_legend=True):
    fig.update_layout(
        height=height,
        margin=dict(l=14, r=14, t=54, b=18),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(
            family='Inter, "Segoe UI", Arial, sans-serif',
            size=12,
            color="#0f172a",
        ),
        hoverlabel=dict(bgcolor="#ffffff", font_size=12),
    )
    if horizontal_legend:
        fig.update_layout(
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1,
            )
        )
    fig.update_xaxes(showgrid=False, zeroline=False)
    fig.update_yaxes(showgrid=True, gridcolor="rgba(148,163,184,0.20)", zeroline=False)
    return fig

def render_plotly(fig, key, filename):
    config = get_plotly_download_config(filename)
    try:
        st.plotly_chart(fig, width="stretch", config=config, theme=None, key=key)
    except TypeError:
        st.plotly_chart(fig, use_container_width=True, config=config, theme=None, key=key)

inject_custom_css()

# ==========================================
# 3. KONEKSI KE DATABASE SUPABASE
# ==========================================
try:
    url: str = st.secrets["SUPABASE_URL"]
    key: str = st.secrets["SUPABASE_KEY"]
    supabase: Client = create_client(url, key)
except Exception:
    st.error("⚠️ Gagal terhubung ke Database. Pastikan SUPABASE_URL dan SUPABASE_KEY sudah ada di Streamlit Secrets!")
    st.stop()

@st.cache_data(ttl=1000)
def load_mapping_sotk(tahun):
    res = (
        supabase.table("mapping_sotk")
        .select("kode_lama, kode_baru")
        .eq("tahun", tahun)
        .execute()
    )
    return {row["kode_lama"]: row["kode_baru"] for row in res.data}

# ==========================================
# 4. SISTEM LOGIN
# ==========================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    render_page_hero(
        "🔒 Akses Terbatas",
        "Aplikasi ini bersifat privat. Masukkan kata sandi untuk melanjutkan.",
    )
    with st.form("login_form"):
        password_input = st.text_input("Kata Sandi", type="password")
        tombol_masuk = st.form_submit_button("Masuk")
        if tombol_masuk:
            if password_input == st.secrets["APP_PASSWORD"]:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Kata sandi salah!")
    st.stop()

# ==========================================
# 5. MENU NAVIGASI MODERN (SIDEBAR)
# ==========================================
with st.sidebar:
    st.markdown(
        """
        <div style="text-align:center; padding: .35rem 0 .2rem 0;">
            <div style="font-size:1.55rem; font-weight:800;">📊 Mamayo Data</div>
            <div style="opacity:.78; font-size:.88rem;">Dashboard olah data SIPD modern</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")

    menu_pilihan = option_menu(
        menu_title=None,
        options=["Alat Excel", "Import SIPD", "Rekap SIPD"],
        icons=["wrench-adjustable", "cloud-arrow-up-fill", "bar-chart-steps"],
        default_index=0,
        key="menu_utama",
        styles={
            "container": {"padding": "0!important", "background-color": "transparent"},
            "icon": {"color": "#7dd3fc", "font-size": "18px"},
            "nav-link": {
                "font-size": "15px",
                "text-align": "left",
                "margin": "6px 0",
                "border-radius": "14px",
                "--hover-color": "rgba(14, 165, 233, 0.18)",
                "padding": "10px 12px",
            },
            "nav-link-selected": {
                "background": "linear-gradient(135deg, #2563eb, #0ea5e9)",
                "color": "white",
                "font-weight": "700",
            },
        },
    )

    st.markdown("---")
    st.caption("🚀 Dikembangkan dengan Python, Streamlit, Plotly, dan Supabase")

# ==========================================
# FUNGSI-FUNGSI UTILITAS REUSABLE
# ==========================================
def terapkan_translasi_sotk(df, mapping_sotk):
    """
    Mesin Translasi SOTK: Mengganti kode_skpd lama -> baru
    agar data dari 2 OPD yang berbeda nama tergabung jadi satu.
    """
    if not mapping_sotk:
        return df
    df = df.copy()
    df["kode_skpd"] = df["kode_skpd"].replace(mapping_sotk)
    return df

def bangun_hierarki(
    df_input,
    list_tahapan_kolom,
    tahap_awal,
    tahap_akhir,
    tahapan_acuan=None,
    df_link=None,
    mode="hierarki",
):
    """
    Fungsi universal pembangun hierarki 5 level.
    Menggantikan duplikasi kode di Tab 1 dan Tab 3.
    """

    def hitung_level(df_src, list_group, level_num):
        if mode == "dpa":
            df_filter = df_src[df_src["tahapan"].isin([tahap_awal, tahap_akhir])]
        else:
            df_filter = df_src.copy()

        grouped = (
            df_filter.groupby(list_group + ["tahapan"])["pagu"]
            .sum()
            .reset_index()
        )
        pivot = (
            grouped.pivot_table(
                index=list_group,
                columns="tahapan",
                values="pagu",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
        )
        pivot["Level"] = level_num
        return pivot

    kumpulan_level = []

    l1 = hitung_level(df_input, ["kode_skpd", "nama_skpd"], 1)
    l1["Kode"], l1["Uraian"], l1["Sort_Key"] = l1["kode_skpd"], l1["nama_skpd"], l1["kode_skpd"]
    kumpulan_level.append(l1)

    df_l2 = df_input[df_input["kode_urusan"] != ""]
    if not df_l2.empty:
        l2 = hitung_level(df_l2, ["kode_skpd", "kode_urusan", "nama_urusan"], 2)
        l2["Kode"], l2["Uraian"] = l2["kode_urusan"], l2["nama_urusan"]
        l2["Sort_Key"] = l2["kode_skpd"] + "|" + l2["kode_urusan"]
        kumpulan_level.append(l2)

    df_l3 = df_input[df_input["kode_program"] != ""]
    if not df_l3.empty:
        l3 = hitung_level(df_l3, ["kode_skpd", "kode_urusan", "kode_program", "nama_program"], 3)
        l3["Kode"], l3["Uraian"] = l3["kode_program"], l3["nama_program"]
        l3["Sort_Key"] = l3["kode_skpd"] + "|" + l3["kode_urusan"] + "|" + l3["kode_program"]
        kumpulan_level.append(l3)

    df_l4 = df_input[df_input["kode_kegiatan"] != ""]
    if not df_l4.empty:
        l4 = hitung_level(
            df_l4,
            ["kode_skpd", "kode_urusan", "kode_program", "kode_kegiatan", "nama_kegiatan"],
            4,
        )
        l4["Kode"], l4["Uraian"] = l4["kode_kegiatan"], l4["nama_kegiatan"]
        l4["Sort_Key"] = (
            l4["kode_skpd"] + "|" + l4["kode_urusan"] + "|" + l4["kode_program"] + "|" + l4["kode_kegiatan"]
        )
        kumpulan_level.append(l4)

    df_l5 = df_input[df_input["kode_sub_kegiatan"] != ""]
    if not df_l5.empty:
        l5 = hitung_level(
            df_l5,
            [
                "kode_skpd",
                "kode_urusan",
                "kode_program",
                "kode_kegiatan",
                "kode_sub_kegiatan",
                "nama_sub_kegiatan",
            ],
            5,
        )
        l5["Kode"], l5["Uraian"] = l5["kode_sub_kegiatan"], l5["nama_sub_kegiatan"]
        l5["Sort_Key"] = (
            l5["kode_skpd"]
            + "|"
            + l5["kode_urusan"]
            + "|"
            + l5["kode_program"]
            + "|"
            + l5["kode_kegiatan"]
            + "|"
            + l5["kode_sub_kegiatan"]
        )

        acuan_sd = tahapan_acuan if tahapan_acuan else tahap_akhir
        df_sd = df_input[df_input["tahapan"] == acuan_sd]
        sd_grouped = (
            df_sd[df_sd["pagu"] > 0]
            .groupby(["kode_sub_kegiatan", "nama_sumber_dana"])["pagu"]
            .sum()
            .reset_index()
        )
        if not sd_grouped.empty:
            sd_grouped["teks_sd"] = (
                sd_grouped["nama_sumber_dana"]
                + " = Rp "
                + sd_grouped["pagu"].apply(lambda x: f"{int(x):,}").str.replace(",", ".")
                + " \n"
            )
            sd_final = (
                sd_grouped.groupby("kode_sub_kegiatan")["teks_sd"]
                .apply(lambda x: "".join(x).strip())
                .reset_index()
            )

            if mode == "dpa":
                sd_final.rename(columns={"teks_sd": "Rincian Sumber Dana"}, inplace=True)
                l5 = pd.merge(l5, sd_final, on="kode_sub_kegiatan", how="left")
            else:
                sd_final.rename(columns={"teks_sd": "Sumber Dana (Acuan)"}, inplace=True)
                l5 = pd.merge(l5, sd_final, on="kode_sub_kegiatan", how="left")

        if df_link is not None and not df_link.empty:
            l5 = pd.merge(l5, df_link, on="kode_sub_kegiatan", how="left")
            l5.rename(columns={"url": "Link DPA"}, inplace=True)

        kumpulan_level.append(l5)

    df_rekap = pd.concat(kumpulan_level, ignore_index=True)

    if mode == "dpa":
        for t in [tahap_awal, tahap_akhir]:
            if t not in df_rekap.columns:
                df_rekap[t] = 0
    else:
        for t in list_tahapan_kolom:
            if t not in df_rekap.columns:
                df_rekap[t] = 0

    df_rekap["Selisih (Akhir - Awal)"] = df_rekap[tahap_akhir] - df_rekap[tahap_awal]

    for col_opsional in ["Sumber Dana (Acuan)", "Rincian Sumber Dana", "Link DPA"]:
        if col_opsional in df_rekap.columns:
            df_rekap[col_opsional] = df_rekap[col_opsional].fillna("")

    df_rekap = df_rekap.sort_values("Sort_Key").reset_index(drop=True)
    return df_rekap

# ==========================================
# 6. KONTEN BERDASARKAN MENU YANG DIPILIH
# ==========================================

# -------------------------------------------------------------------------
# MODUL 1: ALAT EXCEL
# -------------------------------------------------------------------------
if menu_pilihan == "Alat Excel":
    render_page_hero(
        "🛠️ Alat Excel",
        "Rapikan data Dapodik/SIPD dalam satu kali proses dengan tampilan yang lebih modern.",
    )

    file_excel = st.file_uploader("📥 Unggah File Excel (.xlsx)", type=["xlsx"], key="excel_upload")
    if file_excel:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 1️⃣ Pengaturan Tanda Petik")
            kolom_petik = st.text_input("🔠 Kolom Petik (Contoh: C, D)").upper()
            mode_excel = st.radio(
                "⚙️ Aksi",
                ["+ Tambah Petik Tersembunyi", "- Hapus Semua Petik"],
                horizontal=True,
            )
        with col2:
            st.markdown("#### 2️⃣ Pengaturan Pembersih Karakter")
            kolom_bersih = st.text_input("🧹 Kolom Ekstrak Angka (Contoh: F, G)").upper()

        if st.button("🚀 PROSES FILE EXCEL", type="primary", use_container_width=True):
            if not kolom_petik and not kolom_bersih:
                st.error("⚠️ Mohon isi minimal salah satu kolom!")
            else:
                with st.spinner("Memproses data..."):
                    try:
                        list_petik = [k.strip() for k in kolom_petik.split(",") if k.strip()]
                        list_bersih = [k.strip() for k in kolom_bersih.split(",") if k.strip()]

                        wb = openpyxl.load_workbook(file_excel)
                        ws = wb.active

                        if list_bersih:
                            for col in list_bersih:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col}{row}"]
                                    if cell.value is not None:
                                        val_str = re.sub(r"\D", "", str(cell.value).strip())
                                        cell.value = val_str

                        if list_petik:
                            for col in list_petik:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col}{row}"]
                                    if cell.value is not None:
                                        val_str = str(cell.value).strip()
                                        if mode_excel == "+ Tambah Petik Tersembunyi":
                                            val_str = val_str[1:] if val_str.startswith("'") else val_str
                                            cell.value = val_str
                                            cell.quotePrefix = True
                                        else:
                                            cell.value = val_str.replace("'", "")
                                            cell.quotePrefix = False
                                            cell.number_format = "@"

                        output_excel = io.BytesIO()
                        wb.save(output_excel)
                        output_excel.seek(0)

                        st.success("✅ File berhasil diproses!")
                        st.download_button(
                            label="📥 Download Hasil Excel",
                            data=output_excel,
                            file_name=f"Selesai_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_excel.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                        )
                    except Exception as e:
                        st.error(f"❌ Terjadi kesalahan: {e}")

# -------------------------------------------------------------------------
# MODUL 2: IMPORT DATA SIPD & MANAJEMEN DATABASE
# -------------------------------------------------------------------------
elif menu_pilihan == "Import SIPD":
    render_page_hero(
        "📥 Import & Manajemen Data SIPD",
        "Unggah file DPA/RKA (Excel/CSV) atau file backup ke dalam database Supabase.",
    )

    st.markdown("### ➕ Tambah Data Baru")
    with st.form("form_import"):
        st.info("💡 Tahun anggaran ditarik otomatis dari file. Anda cukup mengetik nama tahapan.")
        tahapan_input = st.text_input("🏷️ Nama Tahapan", placeholder="Contoh: Pergeseran 3, Murni, dll")
        file_upload = st.file_uploader("📂 Pilih File Excel / CSV (termasuk file Backup)", type=["xlsx", "xls", "csv"])
        submit_import = st.form_submit_button("🚀 Upload & Simpan ke Database")

    if submit_import:
        if not tahapan_input:
            st.error("❌ Nama Tahapan wajib diisi!")
        elif file_upload is None:
            st.error("❌ File Excel/CSV belum dimasukkan!")
        else:
            with st.spinner("Membaca file dan menyinkronkan dengan database..."):
                try:
                    if file_upload.name.endswith(".csv"):
                        df = pd.read_csv(file_upload)
                    else:
                        df = pd.read_excel(file_upload)

                    df.columns = df.columns.astype(str).str.strip()

                    pemetaan_kolom = {
                        "NO": "no_urut",
                        "TAHUN": "tahun",
                        "KODE URUSAN": "kode_urusan",
                        "NAMA URUSAN": "nama_urusan",
                        "KODE SKPD": "kode_skpd",
                        "NAMA SKPD": "nama_skpd",
                        "KODE SUB UNIT": "kode_sub_unit",
                        "NAMA SUB UNIT": "nama_sub_unit",
                        "KODE BIDANG URUSAN": "kode_bidang_urusan",
                        "NAMA BIDANG URUSAN": "nama_bidang_urusan",
                        "KODE PROGRAM": "kode_program",
                        "NAMA PROGRAM": "nama_program",
                        "KODE KEGIATAN": "kode_kegiatan",
                        "NAMA KEGIATAN": "nama_kegiatan",
                        "KODE SUB KEGIATAN": "kode_sub_kegiatan",
                        "NAMA SUB KEGIATAN": "nama_sub_kegiatan",
                        "KODE SUMBER DANA": "kode_sumber_dana",
                        "NAMA SUMBER DANA": "nama_sumber_dana",
                        "KODE REKENING": "kode_rekening",
                        "NAMA REKENING": "nama_rekening",
                        "PAKET/KELOMPOK": "paket_kelompok",
                        "NAMA PAKET/KELOMPOK": "nama_paket_kelompok",
                        "PAGU": "pagu",
                    }
                    df.rename(columns=pemetaan_kolom, inplace=True)
                    df.columns = df.columns.str.lower()

                    if "tahun" not in df.columns:
                        st.error("❌ Gagal: Kolom 'TAHUN' tidak ditemukan di file.")
                    else:
                        if "id" in df.columns:
                            df = df.drop(columns=["id"])
                        if "created_at" in df.columns:
                            df = df.drop(columns=["created_at"])

                        df["tahapan"] = tahapan_input
                        df = df.astype(object).where(pd.notnull(df), None)

                        data_insert = df.to_dict(orient="records")
                        batch_size = 1000
                        for i in range(0, len(data_insert), batch_size):
                            batch = data_insert[i : i + batch_size]
                            supabase.table("rekap_sipd").insert(batch).execute()

                        st.success(f"✅ {len(df)} baris data '{tahapan_input}' berhasil disimpan ke database!")
                        time.sleep(1)
                        st.rerun()
                except Exception as e:
                    st.error(f"❌ Terjadi kesalahan saat memproses file: {e}")

    st.markdown("---")
    st.markdown("### ⚙️ Manajemen Database")

    with st.expander("⚠️ Buka Panel Zona Berbahaya (Backup & Hapus Data)"):
        st.markdown("#### 1. 📥 Backup Seluruh Database")
        if st.button("📦 Buat File Backup CSV", type="primary"):
            with st.spinner("Menarik seluruh data dari server..."):
                semua_data = []
                offset = 0
                limit = 1000

                while True:
                    res = supabase.table("rekap_sipd").select("*").range(offset, offset + limit - 1).execute()
                    if not res.data:
                        break
                    semua_data.extend(res.data)
                    if len(res.data) < limit:
                        break
                    offset += limit

                if semua_data:
                    df_backup = pd.DataFrame(semua_data)
                    csv_data = df_backup.to_csv(index=False).encode("utf-8")
                    st.success(f"✅ Backup siap! Total: {len(df_backup)} baris data.")
                    st.download_button(
                        label="⬇️ Download Backup.csv",
                        data=csv_data,
                        file_name="Backup_Database_SIPD.csv",
                        mime="text/csv",
                    )
                else:
                    st.warning("Database masih kosong. Tidak ada yang bisa di-backup.")

        st.markdown("<hr style='border: 1px dashed #ccc;'>", unsafe_allow_html=True)
        st.markdown("#### 2. 🗑️ Hapus Data Parsial (Sesuai Database)")

        try:
            semua_opsi = []
            offset = 0
            limit = 1000
            while True:
                res_opsi = supabase.table("rekap_sipd").select("tahun, tahapan").range(offset, offset + limit - 1).execute()
                if not res_opsi.data:
                    break
                semua_opsi.extend(res_opsi.data)
                if len(res_opsi.data) < limit:
                    break
                offset += limit

            if semua_opsi:
                df_opsi = pd.DataFrame(semua_opsi)
                unique_years = sorted(df_opsi["tahun"].dropna().unique().tolist())
            else:
                unique_years = []
                df_opsi = pd.DataFrame()
        except Exception:
            unique_years = []
            df_opsi = pd.DataFrame()

        if not unique_years:
            st.info("Database masih kosong, belum ada data yang bisa dihapus.")
        else:
            col_del1, col_del2 = st.columns(2)
            with col_del1:
                del_tahun = st.selectbox("Pilih Tahun", unique_years, key="del_thn")
            with col_del2:
                tahapan_tersedia = sorted(df_opsi[df_opsi["tahun"] == del_tahun]["tahapan"].dropna().unique().tolist())
                del_tahapan = st.selectbox(f"Pilih Tahapan di {del_tahun}", tahapan_tersedia, key="del_thp")

            if del_tahapan:
                if st.button(f"🗑️ Hapus Data {del_tahapan} {del_tahun}"):
                    with st.spinner("Menghapus data..."):
                        supabase.table("rekap_sipd").delete().eq("tahun", del_tahun).eq("tahapan", del_tahapan).execute()
                        st.success(f"✅ Data {del_tahapan} Tahun {del_tahun} berhasil dihapus!")
                        time.sleep(1)
                        st.rerun()

        st.markdown("<hr style='border: 1px dashed #ccc;'>", unsafe_allow_html=True)
        st.markdown("#### 3. 🔥 Factory Reset (Kosongkan Database)")
        konfirmasi_kiamat = st.text_input("Ketik 'HAPUS TOTAL' untuk membuka kunci eksekusi")

        if konfirmasi_kiamat == "HAPUS TOTAL":
            if st.button("🚨 EKSEKUSI TRUNCATE & RESET ID", type="primary"):
                with st.spinner("Menghancurkan seluruh data..."):
                    try:
                        supabase.rpc("truncate_rekap_sipd").execute()
                        st.success("💥 Database berhasil dikosongkan total.")
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Gagal: {e}. Pastikan fungsi RPC sudah dibuat di Supabase.")

# -------------------------------------------------------------------------
# MODUL 3: REKAP SIPD
# -------------------------------------------------------------------------
elif menu_pilihan == "Rekap SIPD":
    render_page_hero(
        "📊 Rekapitulasi SIPD",
        "Analisis postur anggaran antar tahapan, dashboard modern, dan integrasi DPA.",
    )

    @st.cache_data(ttl=3600, show_spinner=False)
    def tarik_data_database():
        semua_data = []
        offset = 0
        limit = 1000
        while True:
            res = supabase.table("rekap_sipd").select("*").order("id").range(offset, offset + limit - 1).execute()
            data_tarikan = res.data
            if not data_tarikan:
                break
            semua_data.extend(data_tarikan)
            if len(data_tarikan) < limit:
                break
            offset += limit
        return pd.DataFrame(semua_data)

    with st.spinner("⏳ Menarik seluruh data dari database..."):
        try:
            df_mentah = tarik_data_database()
        except Exception as e:
            st.error(f"❌ Gagal menarik data: {e}")
            df_mentah = pd.DataFrame()

    if st.button("🔄 Refresh Data Database"):
        tarik_data_database.clear()
        st.rerun()

    if df_mentah.empty:
        st.info("💡 Database masih kosong. Silakan Import SIPD terlebih dahulu.")
        st.stop()

    # ==========================================
    # 1. PERSIAPAN DATA DASAR
    # ==========================================
    df = df_mentah.copy()
    df["pagu"] = pd.to_numeric(df["pagu"], errors="coerce").fillna(0)

    if "tahun" in df.columns:
        df["tahun"] = df["tahun"].astype(str).str.replace(".0", "", regex=False)
    else:
        st.error("⚠️ Kolom 'tahun' tidak ditemukan!")
        st.stop()

    df = df.fillna("")
    st.markdown("### ⚙️ Pengaturan Filter & Parameter")

    # ==========================================
    # 2. SISTEM FILTER GLOBAL
    # ==========================================
    list_tahun = sorted([t for t in df["tahun"].unique() if t != ""], reverse=True)
    col_thn, col_skpd = st.columns(2)
    with col_thn:
        tahun_pilihan = st.selectbox("📅 Pilih Tahun Anggaran", list_tahun)

    df_tahun = df[df["tahun"] == tahun_pilihan].copy()

    # FIX: reload mapping saat tahun berubah
    if st.session_state.get("mapping_tahun_aktif") != tahun_pilihan:
        st.session_state.mapping_sotk = load_mapping_sotk(tahun_pilihan)
        st.session_state["mapping_tahun_aktif"] = tahun_pilihan

    list_skpd = sorted([s for s in df_tahun["nama_skpd"].unique() if s != ""])
    list_skpd.insert(0, "SEMUA SKPD")

    with col_skpd:
        skpd_pilihan = st.multiselect(
            "🏢 Pilih SKPD (bisa pilih lebih dari 1 untuk digabung)",
            list_skpd,
            default=["SEMUA SKPD"],
        )

    if "SEMUA SKPD" in skpd_pilihan:
        df_proses = df_tahun.copy()
        nama_file_export = "SEMUA_SKPD"
    else:
        df_proses = df_tahun[df_tahun["nama_skpd"].isin(skpd_pilihan)].copy()
        nama_file_export = "MERGER_" + "_".join([s.replace(" ", "")[:10] for s in skpd_pilihan])

    if df_proses.empty:
        st.warning(f"⚠️ Tidak ada data untuk pilihan SKPD tersebut di tahun {tahun_pilihan}.")
        st.stop()

    tahapan_tersedia = [t for t in df_proses["tahapan"].unique() if t != ""]

    st.markdown("#### 📋 Urutan Tahapan & Acuan Selisih")
    list_tahapan = st.multiselect(
        "Susun urutan tahapan (kiri ke kanan)",
        options=tahapan_tersedia,
        default=tahapan_tersedia,
    )

    if not list_tahapan:
        st.error("⚠️ Pilih minimal 1 tahapan.")
        st.stop()

    col_t1, col_t2, col_t3 = st.columns(3)
    with col_t1:
        tahapan_acuan = st.selectbox("🎯 Acuan Sumber Dana (khusus Tab 1)", list_tahapan, index=len(list_tahapan) - 1)
    with col_t2:
        tahap_awal = st.selectbox("📉 Tahapan Awal (pengurang)", list_tahapan, index=0)
    with col_t3:
        tahap_akhir = st.selectbox("📈 Tahapan Akhir (dikurangi)", list_tahapan, index=len(list_tahapan) - 1)

    # ==========================================
    # 2.5 MAPPING PERUBAHAN SOTK
    # ==========================================
    if st.session_state.get("mapping_alert") == "sukses":
        st.success("✅ Mapping SOTK sudah disimpan ke database!")
        st.session_state["mapping_alert"] = None
    if st.session_state.get("mapping_alert") == "hapus":
        st.success("✅ Mapping berhasil dihapus!")
        st.session_state["mapping_alert"] = None
    if st.session_state.get("mapping_alert") == "hapus_semua":
        st.success("✅ Semua mapping berhasil dihapus!")
        st.session_state["mapping_alert"] = None

    with st.expander("🔄 Mapping Perubahan SOTK / Perubahan Nama OPD (Opsional)", expanded=False):
        st.caption("Tambahkan mapping agar data OPD lama tergabung dengan OPD baru.")
        df_skpd_unik = df_proses[["kode_skpd", "nama_skpd"]].drop_duplicates().sort_values("kode_skpd")
        daftar_opsi_skpd = [f"{row['kode_skpd']}  |  {row['nama_skpd']}" for _, row in df_skpd_unik.iterrows()]

        col_lama, col_baru, col_btn = st.columns([4, 4, 2])
        with col_lama:
            opd_lama_pilihan = st.selectbox("OPD LAMA", daftar_opsi_skpd, key="opd_lama_sel")
        with col_baru:
            opd_baru_pilihan = st.selectbox("OPD BARU", daftar_opsi_skpd, key="opd_baru_sel")
        with col_btn:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("➕ Tambah", type="primary", key="btn_tambah_sotk_db", use_container_width=True):
                kode_lama = opd_lama_pilihan.split("  |  ")[0].strip()
                kode_baru = opd_baru_pilihan.split("  |  ")[0].strip()
                if kode_lama == kode_baru:
                    st.error("❌ OPD Lama dan Baru tidak boleh sama!")
                else:
                    supabase.table("mapping_sotk").insert(
                        [
                            {
                                "kode_lama": kode_lama,
                                "kode_baru": kode_baru,
                                "tahun": tahun_pilihan,
                                "username": st.session_state.get("username", ""),
                            }
                        ]
                    ).execute()
                    load_mapping_sotk.clear()
                    st.session_state["mapping_alert"] = "sukses"
                    st.session_state.mapping_sotk = load_mapping_sotk(tahun_pilihan)
                    st.rerun()

        if st.session_state.get("mapping_sotk"):
            st.markdown("##### 📋 Mapping SOTK Aktif (Database)")
            for idx, (k_lama, k_baru) in enumerate(st.session_state.mapping_sotk.items()):
                col_info, col_hapus = st.columns([8, 2])
                with col_info:
                    st.markdown(f"🔸 `{k_lama}` → `{k_baru}`")
                with col_hapus:
                    if st.button("🗑️ Hapus", key=f"hapus_sotk_{k_lama}_{k_baru}_{idx}"):
                        resp = supabase.table("mapping_sotk").delete().eq("kode_lama", k_lama).eq("tahun", tahun_pilihan).execute()
                        if resp.data:
                            load_mapping_sotk.clear()
                            st.session_state["mapping_alert"] = "hapus"
                            st.session_state.mapping_sotk = load_mapping_sotk(tahun_pilihan)
                            st.rerun()
                        else:
                            st.error("❌ Gagal menghapus mapping di database Supabase.")

            if st.button("🧹 Hapus Semua Mapping", key="hapus_semua_sotk_db"):
                resp = supabase.table("mapping_sotk").delete().eq("tahun", tahun_pilihan).execute()
                if resp.data is not None:
                    load_mapping_sotk.clear()
                    st.session_state["mapping_alert"] = "hapus_semua"
                    st.session_state.mapping_sotk = {}
                    st.rerun()
                else:
                    st.error("❌ Gagal menghapus semua mapping di database Supabase.")
        else:
            st.info("Belum ada mapping. Sistem akan berfungsi seperti biasa.")

    # ==========================================
    # 2.6 TERAPKAN TRANSLASI SOTK & JANGKAR NOMENKLATUR
    # ==========================================
    df_proses = terapkan_translasi_sotk(df_proses, st.session_state.get("mapping_sotk", {}))

    df_akhir = df_proses[df_proses["tahapan"] == tahap_akhir]
    if not df_akhir.empty:
        kolom_hierarki = [
            "kode_skpd",
            "nama_skpd",
            "kode_urusan",
            "nama_urusan",
            "kode_program",
            "nama_program",
            "kode_kegiatan",
            "nama_kegiatan",
            "nama_sub_kegiatan",
        ]

        df_ref = (
            df_akhir[["kode_sub_kegiatan"] + kolom_hierarki]
            .drop_duplicates("kode_sub_kegiatan")
            .set_index("kode_sub_kegiatan")
        )

        for col in kolom_hierarki:
            dict_map = df_ref[col].to_dict()
            df_proses[col] = df_proses["kode_sub_kegiatan"].map(dict_map).fillna(df_proses[col])

        df_ref_skpd = df_akhir[["kode_skpd", "nama_skpd"]].drop_duplicates("kode_skpd").set_index("kode_skpd")
        dict_nama_skpd_akhir = df_ref_skpd["nama_skpd"].to_dict()
        df_proses["nama_skpd"] = df_proses["kode_skpd"].map(dict_nama_skpd_akhir).fillna(df_proses["nama_skpd"])

    # ==========================================
    # 3. TAB MENU
    # ==========================================
    tab0, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
        [
            "📊 Dashboard",
            "📑 Rekap Hierarki",
            "💰 Rekap Sumber Dana",
            "🔗 Integrasi Link DPA",
            "📈 Evaluasi Realisasi",
            "🏢 Rekap Per Bidang",
            "📦 Rekap Kode Rekening",
        ]
    )

    # -------------------------------------------------------------------
    # TAB 0: DASHBOARD
    # -------------------------------------------------------------------
    with tab0:
        st.markdown(
            f"""
            <div class="hero-box">
                <h3>Dashboard Anggaran {tahun_pilihan}</h3>
                <p>
                    Perbandingan <b>{tahap_awal}</b> vs <b>{tahap_akhir}</b> ·
                    {len(list_tahapan)} tahapan aktif ·
                    Mode data: <b>{nama_file_export.replace("_", " ")}</b>
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.caption("Tip: gunakan ikon kamera di kanan atas tiap chart untuk mengunduh PNG dengan resolusi lebih tajam.")

        df_dash = df_proses.copy()
        metrik_per_tahapan = df_dash.groupby("tahapan")["pagu"].sum()

        pagu_awal = metrik_per_tahapan.get(tahap_awal, 0)
        pagu_akhir = metrik_per_tahapan.get(tahap_akhir, 0)
        selisih_total = pagu_akhir - pagu_awal

        jumlah_skpd = df_dash[df_dash["tahapan"] == tahap_akhir]["kode_skpd"].nunique()
        jumlah_sub_keg = df_dash[df_dash["tahapan"] == tahap_akhir]["kode_sub_kegiatan"].nunique()
        jumlah_program = df_dash[df_dash["tahapan"] == tahap_akhir]["kode_program"].nunique()

        m1, m2, m3, m4 = st.columns(4)
        with m1:
            st.metric("💰 Total Pagu Akhir", format_rupiah(pagu_akhir))
        with m2:
            st.metric("📉 Total Pagu Awal", format_rupiah(pagu_awal))
        with m3:
            st.metric("🏢 SKPD / Program", f"{jumlah_skpd} / {jumlah_program}")
        with m4:
            st.metric(
                "📊 Selisih Total",
                format_rupiah(abs(selisih_total)),
                delta=f"{'+' if selisih_total >= 0 else '-'} {format_rupiah(abs(selisih_total))}",
            )

        st.caption(f"Sub kegiatan aktif pada {tahap_akhir}: {jumlah_sub_keg}")

        col_chart1, col_chart2 = st.columns([1.15, 0.85], gap="large")

        with col_chart1:
            df_bar = pd.DataFrame(
                {
                    "Tahapan": list_tahapan,
                    "Total Pagu": [metrik_per_tahapan.get(t, 0) for t in list_tahapan],
                }
            )
            df_bar["Label"] = df_bar["Total Pagu"].apply(format_rupiah)

            fig_bar = px.bar(
                df_bar,
                x="Tahapan",
                y="Total Pagu",
                text="Label",
                color="Tahapan",
                color_discrete_sequence=CHART_PALETTE,
            )
            fig_bar.update_traces(
                textposition="outside",
                cliponaxis=False,
                customdata=df_bar[["Label"]],
                hovertemplate="<b>%{x}</b><br>%{customdata[0]}<extra></extra>",
            )
            style_fig(fig_bar, height=430)
            fig_bar.update_layout(title="Perbandingan Total Pagu per Tahapan", xaxis_title="", yaxis_title="")
            render_plotly(
                fig_bar,
                key="dashboard_bar_tahapan",
                filename=safe_stem(f"pagu_tahapan_{tahun_pilihan}"),
            )

        with col_chart2:
            df_sd_dash = df_dash[df_dash["tahapan"] == tahap_akhir].copy()
            df_sd_dash["nama_sumber_dana"] = df_sd_dash["nama_sumber_dana"].replace("", "TIDAK DIKETAHUI")
            sd_pie = (
                df_sd_dash.groupby("nama_sumber_dana")["pagu"]
                .sum()
                .reset_index()
                .query("pagu > 0")
                .sort_values("pagu", ascending=False)
            )

            if not sd_pie.empty:
                if len(sd_pie) > 6:
                    top_sd = sd_pie.head(5).copy()
                    lainnya = sd_pie.iloc[5:]["pagu"].sum()
                    if lainnya > 0:
                        top_sd = pd.concat(
                            [
                                top_sd,
                                pd.DataFrame([{"nama_sumber_dana": "Lainnya", "pagu": lainnya}]),
                            ],
                            ignore_index=True,
                        )
                    sd_pie_show = top_sd
                else:
                    sd_pie_show = sd_pie.copy()

                sd_pie_show["Label"] = sd_pie_show["pagu"].apply(format_rupiah)

                fig_pie = px.pie(
                    sd_pie_show,
                    names="nama_sumber_dana",
                    values="pagu",
                    hole=0.62,
                    color_discrete_sequence=CHART_PALETTE[:6],
                )
                fig_pie.update_traces(
                    textinfo="percent",
                    customdata=sd_pie_show[["Label"]],
                    hovertemplate="<b>%{label}</b><br>%{customdata[0]}<br>%{percent}<extra></extra>",
                )
                fig_pie.add_annotation(
                    text=f"<b>Total</b><br>{format_rupiah(sd_pie_show['pagu'].sum())}",
                    x=0.5,
                    y=0.5,
                    showarrow=False,
                )
                style_fig(fig_pie, height=430)
                fig_pie.update_layout(title=f"Komposisi Sumber Dana · {tahap_akhir}")
                render_plotly(
                    fig_pie,
                    key="dashboard_pie_sd",
                    filename=safe_stem(f"sumber_dana_{tahun_pilihan}_{tahap_akhir}"),
                )
            else:
                st.info("Tidak ada data sumber dana untuk ditampilkan.")

        df_skpd_dash = (
            df_dash[df_dash["tahapan"] == tahap_akhir]
            .groupby(["kode_skpd", "nama_skpd"])["pagu"]
            .sum()
            .reset_index()
            .sort_values("pagu", ascending=False)
        )

        if not df_skpd_dash.empty:
            total_skpd_di_chart = len(df_skpd_dash)
            df_skpd_show = df_skpd_dash.head(15).copy()
            df_skpd_show = df_skpd_show.sort_values("pagu", ascending=True)
            df_skpd_show["label_skpd"] = df_skpd_show["nama_skpd"].apply(lambda x: shorten_text(x, 42))
            df_skpd_show["Label"] = df_skpd_show["pagu"].apply(format_rupiah)

            fig_skpd = px.bar(
                df_skpd_show,
                x="pagu",
                y="label_skpd",
                orientation="h",
                text="Label",
            )
            fig_skpd.update_traces(
                marker_color=COLOR_PRIMARY,
                textposition="outside",
                cliponaxis=False,
                customdata=df_skpd_show[["Label", "nama_skpd"]],
                hovertemplate="<b>%{customdata[1]}</b><br>%{customdata[0]}<extra></extra>",
            )
            style_fig(fig_skpd, height=max(430, len(df_skpd_show) * 38), horizontal_legend=False)
            fig_skpd.update_layout(title=f"15 SKPD dengan Pagu Terbesar · {tahap_akhir}", xaxis_title="", yaxis_title="")
            render_plotly(
                fig_skpd,
                key="dashboard_bar_skpd",
                filename=safe_stem(f"top_skpd_{tahun_pilihan}_{tahap_akhir}"),
            )

            if total_skpd_di_chart > 15:
                st.caption(f"Menampilkan 15 dari {total_skpd_di_chart} SKPD.")
        else:
            st.info("Tidak ada data SKPD untuk grafik.")

        st.markdown("#### 🔝 Top 10 Sub Kegiatan dengan Perubahan Terbesar")

        df_selisih = (
            df_dash.groupby(["kode_sub_kegiatan", "nama_sub_kegiatan", "nama_skpd", "tahapan"])["pagu"]
            .sum()
            .reset_index()
        )
        pivot_selisih = (
            df_selisih.pivot_table(
                index=["kode_sub_kegiatan", "nama_sub_kegiatan", "nama_skpd"],
                columns="tahapan",
                values="pagu",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
        )

        for t in [tahap_awal, tahap_akhir]:
            if t not in pivot_selisih.columns:
                pivot_selisih[t] = 0

        pivot_selisih["Selisih"] = pivot_selisih[tahap_akhir] - pivot_selisih[tahap_awal]
        pivot_selisih["Abs_Selisih"] = pivot_selisih["Selisih"].abs()
        top10 = pivot_selisih.nlargest(10, "Abs_Selisih").copy()

        if not top10.empty:
            top10["Arah"] = top10["Selisih"].apply(lambda x: "Naik" if x >= 0 else "Turun")
            top10["label_sub"] = top10["nama_sub_kegiatan"].apply(lambda x: shorten_text(x, 48))
            top10["SelisihLabel"] = top10["Selisih"].apply(
                lambda x: f"{'+' if x >= 0 else '-'} {format_rupiah(abs(x))}"
            )

            top10_chart = top10.sort_values("Selisih").copy()

            fig_delta = px.bar(
                top10_chart,
                x="Selisih",
                y="label_sub",
                orientation="h",
                color="Arah",
                color_discrete_map={"Naik": COLOR_SUCCESS, "Turun": COLOR_DANGER},
            )
            fig_delta.update_traces(
                customdata=top10_chart[["SelisihLabel", "nama_skpd"]],
                hovertemplate="<b>%{y}</b><br>%{customdata[0]}<br>SKPD: %{customdata[1]}<extra></extra>",
            )
            style_fig(fig_delta, height=500)
            fig_delta.update_layout(title=f"Perubahan {tahap_akhir} vs {tahap_awal}", xaxis_title="", yaxis_title="")
            render_plotly(
                fig_delta,
                key="dashboard_delta_subkeg",
                filename=safe_stem(f"delta_sub_kegiatan_{tahun_pilihan}"),
            )

            nama_awal = f"Pagu {tahap_awal}"
            nama_akhir = f"Pagu {tahap_akhir}"
            if nama_awal == nama_akhir:
                nama_awal = f"Pagu {tahap_awal} (Awal)"
                nama_akhir = f"Pagu {tahap_akhir} (Akhir)"

            tabel_top10 = top10[
                ["kode_sub_kegiatan", "nama_sub_kegiatan", "nama_skpd", tahap_awal, tahap_akhir, "Selisih"]
            ].copy()
            tabel_top10.columns = ["Kode Sub", "Uraian Sub Kegiatan", "SKPD", nama_awal, nama_akhir, "Selisih"]

            with st.expander("Lihat tabel detail Top 10", expanded=True):
                st.dataframe(
                    tabel_top10,
                    use_container_width=True,
                    column_config={
                        nama_awal: st.column_config.NumberColumn(format="Rp %.0f"),
                        nama_akhir: st.column_config.NumberColumn(format="Rp %.0f"),
                        "Selisih": st.column_config.NumberColumn(format="Rp %.0f"),
                    },
                )
        else:
            st.info("Tidak cukup data untuk menampilkan perubahan sub kegiatan.")

        st.markdown("#### 🔍 Cari Sub Kegiatan, Kode Sub, atau SKPD")
        keyword = st.text_input(
            "Kata kunci pencarian",
            placeholder="Contoh: sanitasi, 1.02.03, atau nama SKPD",
        )

        if keyword:
            mask = (
                df_dash["nama_sub_kegiatan"].str.contains(keyword, case=False, na=False)
                | df_dash["kode_sub_kegiatan"].str.contains(keyword, case=False, na=False)
                | df_dash["nama_skpd"].str.contains(keyword, case=False, na=False)
            )
            hasil = df_dash[mask].copy()
            st.success(f"Menemukan {len(hasil)} baris hasil pencarian.")

            kolom_tampil = [
                "kode_sub_kegiatan",
                "nama_sub_kegiatan",
                "nama_skpd",
                "tahapan",
                "pagu",
                "nama_sumber_dana",
            ]
            st.dataframe(
                hasil[kolom_tampil].sort_values("pagu", ascending=False),
                use_container_width=True,
                column_config={"pagu": st.column_config.NumberColumn(format="Rp %.0f")},
            )

            if len(hasil) > 0:
                hasil_csv = hasil[kolom_tampil].to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="📥 Download Hasil Pencarian (CSV)",
                    data=hasil_csv,
                    file_name="Hasil_Pencarian_SubKegiatan.csv",
                    mime="text/csv",
                )
        else:
            st.info("Masukkan kata kunci untuk mulai mencari.")

    # -------------------------------------------------------------------
    # TAB 1: REKAP HIERARKI
    # -------------------------------------------------------------------
    with tab1:
        if st.button("🚀 PROSES LAPORAN HIERARKI", type="primary", use_container_width=True, key="btn_tab1"):
            with st.spinner("Memproses Laporan Hierarki..."):
                df_rekap = bangun_hierarki(
                    df_input=df_proses,
                    list_tahapan_kolom=list_tahapan,
                    tahap_awal=tahap_awal,
                    tahap_akhir=tahap_akhir,
                    tahapan_acuan=tahapan_acuan,
                    mode="hierarki",
                )

                if "Sumber Dana (Acuan)" not in df_rekap.columns:
                    df_rekap["Sumber Dana (Acuan)"] = ""

                kolom_final = ["Kode", "Uraian", "Sumber Dana (Acuan)", "Level"] + list_tahapan + ["Selisih (Akhir - Awal)"]
                df_hasil = df_rekap[[c for c in kolom_final if c in df_rekap.columns]]

                df_tampil = df_hasil.drop(columns=["Level"])
                kolom_angka = list_tahapan + ["Selisih (Akhir - Awal)"]
                format_dict = {col: "{:,.0f}" for col in kolom_angka if col in df_tampil.columns}
                styled_df_web = df_tampil.style.format(format_dict).set_properties(
                    subset=["Sumber Dana (Acuan)"], **{"white-space": "pre-wrap"}
                )

                st.success("✅ Laporan Hierarki berhasil dibuat!")
                st.dataframe(styled_df_web, use_container_width=True, height=500)

                def warna_baris_excel(row):
                    lvl = df_hasil.loc[row.name, "Level"]
                    if lvl == 1:
                        return ["background-color: #ddebf7; font-weight: bold"] * len(row)
                    if lvl == 2:
                        return ["background-color: #fff2cc; font-weight: bold"] * len(row)
                    if lvl == 3:
                        return ["background-color: #fce4d6; font-weight: bold"] * len(row)
                    if lvl == 4:
                        return ["background-color: #e2efda; font-weight: bold"] * len(row)
                    return [""] * len(row)

                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
                    df_tampil.style.apply(warna_baris_excel, axis=1).format(format_dict).to_excel(
                        writer, index=False, sheet_name="Hierarki"
                    )
                output_excel.seek(0)

                st.download_button(
                    "📥 Download Excel (Hierarki)",
                    output_excel,
                    f"Hierarki_{nama_file_export}_{tahun_pilihan}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t1",
                )

    # -------------------------------------------------------------------
    # TAB 2: REKAP SUMBER DANA
    # -------------------------------------------------------------------
    with tab2:
        if st.button("🚀 PROSES REKAP SUMBER DANA", type="primary", use_container_width=True, key="btn_tab2"):
            with st.spinner("Menghitung pagu per sumber dana..."):
                df_sd = df_proses.copy()
                df_sd["nama_sumber_dana"] = df_sd["nama_sumber_dana"].replace("", "TIDAK DIKETAHUI / KOSONG")
                rekap_sd = df_sd.groupby(["nama_sumber_dana", "tahapan"])["pagu"].sum().unstack(fill_value=0).reset_index()

                for t in list_tahapan:
                    if t not in rekap_sd.columns:
                        rekap_sd[t] = 0

                rekap_sd["Selisih (Akhir - Awal)"] = rekap_sd[tahap_akhir] - rekap_sd[tahap_awal]
                rekap_sd = rekap_sd.sort_values(by=tahap_akhir, ascending=False).reset_index(drop=True)

                kolom_angka_sd = list_tahapan + ["Selisih (Akhir - Awal)"]
                kolom_final_sd = ["nama_sumber_dana"] + kolom_angka_sd
                df_hasil_sd = rekap_sd[kolom_final_sd].copy()

                baris_total = pd.DataFrame([df_hasil_sd[kolom_angka_sd].sum()])
                baris_total["nama_sumber_dana"] = "=== TOTAL KESELURUHAN ==="
                df_hasil_sd = pd.concat([df_hasil_sd, baris_total], ignore_index=True)

                format_dict_sd = {col: "{:,.0f}" for col in kolom_angka_sd}
                st.success("✅ Rekap sumber dana berhasil dibuat!")
                st.dataframe(df_hasil_sd.style.format(format_dict_sd), use_container_width=True, height=500)

                def highlight_total_excel(row):
                    if row["nama_sumber_dana"] == "=== TOTAL KESELURUHAN ===":
                        return ["background-color: #ffe699; font-weight: bold"] * len(row)
                    return [""] * len(row)

                output_excel_sd = io.BytesIO()
                with pd.ExcelWriter(output_excel_sd, engine="openpyxl") as writer:
                    df_hasil_sd.style.apply(highlight_total_excel, axis=1).format(format_dict_sd).to_excel(
                        writer, index=False, sheet_name="SumberDana"
                    )
                output_excel_sd.seek(0)

                st.download_button(
                    "📥 Download Excel (Sumber Dana)",
                    output_excel_sd,
                    f"SumberDana_{nama_file_export}_{tahun_pilihan}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t2",
                )

    # -------------------------------------------------------------------
    # TAB 3: INTEGRASI LINK DPA
    # -------------------------------------------------------------------
    with tab3:
        st.info(f"💡 Menampilkan perbandingan: **{tahap_awal}** vs **{tahap_akhir}**")

        sumber_data_dpa = st.radio(
            "Pilih Mode Input Link DPA",
            ["📂 Upload File Lokal (Excel/CSV)", "🌐 Link Google Sheet (Otomatis Baca Sheet)"],
            horizontal=True,
            key="radio_dpa",
        )

        file_link = None
        link_dpa_input = ""
        df_link_gsheet = pd.DataFrame()

        if sumber_data_dpa == "📂 Upload File Lokal (Excel/CSV)":
            file_link = st.file_uploader(
                "📂 Upload File Excel Link DPA (harus ada kolom 'kode sub' dan 'url')",
                type=["xlsx", "xls", "csv"],
                key="up_link",
            )
        else:
            link_dpa_input = st.text_input(
                "🔗 Paste Link Google Sheet DPA",
                placeholder="https://docs.google.com/spreadsheets/d/...",
            )
            st.caption("Pastikan akses diatur ke: Anyone with the link")

            if link_dpa_input:
                match = re.search(r"/d/([a-zA-Z0-9-_]+)", link_dpa_input)
                if match:
                    doc_id = match.group(1)
                    url_xlsx = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"

                    try:
                        @st.cache_data(show_spinner=False, ttl=600)
                        def tarik_excel_google(url):
                            resp = requests.get(url)
                            resp.raise_for_status()
                            return resp.content

                        with st.spinner("🔍 Membaca Google Sheet dan mencari daftar sheet..."):
                            excel_bytes = tarik_excel_google(url_xlsx)
                            xls = pd.ExcelFile(io.BytesIO(excel_bytes))
                            daftar_sheet = xls.sheet_names

                        if daftar_sheet:
                            sheet_pilihan = st.selectbox("📑 Pilih Tahapan (Sheet)", daftar_sheet)
                            if sheet_pilihan:
                                df_link_gsheet = pd.read_excel(xls, sheet_name=sheet_pilihan)
                        else:
                            st.error("❌ Tidak ada sheet yang ditemukan di file tersebut.")
                    except Exception as e:
                        st.error(f"❌ Gagal membaca Google Sheet. Error: {e}")
                else:
                    st.warning("⚠️ Link tidak valid. Coba paste ulang link Google Sheet yang benar.")

        if st.button("🚀 PROSES & GABUNGKAN LINK DPA", type="primary", use_container_width=True, key="btn_tab3"):
            if sumber_data_dpa == "📂 Upload File Lokal (Excel/CSV)" and file_link is None:
                st.error("⚠️ Mohon upload file Excel/CSV Link DPA terlebih dahulu!")
            elif sumber_data_dpa == "🌐 Link Google Sheet (Otomatis Baca Sheet)" and link_dpa_input == "":
                st.error("⚠️ Mohon paste Link Google Sheet terlebih dahulu!")
            elif sumber_data_dpa == "🌐 Link Google Sheet (Otomatis Baca Sheet)" and df_link_gsheet.empty:
                st.error("⚠️ Menunggu data dari Google Sheet. Silakan pilih sheet yang benar.")
            else:
                with st.spinner("Menjahit Link DPA dengan Data Anggaran..."):
                    try:
                        if sumber_data_dpa == "📂 Upload File Lokal (Excel/CSV)":
                            if file_link.name.endswith(".csv"):
                                df_link = pd.read_csv(file_link)
                            else:
                                df_link = pd.read_excel(file_link)
                        else:
                            df_link = df_link_gsheet.copy()

                        df_link.columns = df_link.columns.astype(str).str.lower().str.strip()

                        if "kode sub" not in df_link.columns or "url" not in df_link.columns:
                            st.error(
                                f"❌ Gagal! File/Sheet upload tidak punya kolom 'kode sub' atau 'url'. "
                                f"Kolom yang terdeteksi: {list(df_link.columns)}"
                            )
                        else:
                            df_link = df_link[["kode sub", "url"]].rename(columns={"kode sub": "kode_sub_kegiatan"})
                            df_link["kode_sub_kegiatan"] = df_link["kode_sub_kegiatan"].astype(str).str.strip()
                            df_link["url"] = df_link["url"].fillna("")

                            df_rekap_dpa = bangun_hierarki(
                                df_input=df_proses,
                                list_tahapan_kolom=[tahap_awal, tahap_akhir],
                                tahap_awal=tahap_awal,
                                tahap_akhir=tahap_akhir,
                                tahapan_acuan=tahap_akhir,
                                df_link=df_link,
                                mode="dpa",
                            )

                            df_rekap_dpa["Anggaran Sebelum"] = df_rekap_dpa[tahap_awal] if tahap_awal in df_rekap_dpa.columns else 0
                            df_rekap_dpa["Anggaran Sesudah"] = df_rekap_dpa[tahap_akhir] if tahap_akhir in df_rekap_dpa.columns else 0
                            df_rekap_dpa["Selisih"] = df_rekap_dpa["Anggaran Sesudah"] - df_rekap_dpa["Anggaran Sebelum"]

                            for col_opt in ["Rincian Sumber Dana", "Link DPA"]:
                                if col_opt not in df_rekap_dpa.columns:
                                    df_rekap_dpa[col_opt] = ""
                                df_rekap_dpa[col_opt] = df_rekap_dpa[col_opt].fillna("")

                            # FIX TYPO: Urian -> Uraian
                            kolom_final_dpa = [
                                "Link DPA",
                                "Kode",
                                "Uraian",
                                "Rincian Sumber Dana",
                                "Anggaran Sebelum",
                                "Anggaran Sesudah",
                                "Selisih",
                                "Level",
                            ]
                            df_hasil_dpa = df_rekap_dpa[[c for c in kolom_final_dpa if c in df_rekap_dpa.columns]].copy()
                            df_tampil_dpa = df_hasil_dpa.drop(columns=["Level"])

                            st.success("✅ Integrasi Link DPA berhasil!")
                            st.dataframe(
                                df_tampil_dpa,
                                use_container_width=True,
                                height=500,
                                column_config={
                                    "Link DPA": st.column_config.LinkColumn("Link DPA", display_text="🔗 Buka DPA"),
                                    "Anggaran Sebelum": st.column_config.NumberColumn(format="%.0f"),
                                    "Anggaran Sesudah": st.column_config.NumberColumn(format="%.0f"),
                                    "Selisih": st.column_config.NumberColumn(format="%.0f"),
                                },
                            )

                            def format_excel_dpa(row):
                                if pd.notna(row["Link DPA"]) and str(row["Link DPA"]).startswith("http"):
                                    row["Link DPA"] = f'=HYPERLINK("{row["Link DPA"]}", "🔗 Buka DPA")'
                                else:
                                    row["Link DPA"] = ""
                                return row

                            df_excel_dpa = df_hasil_dpa.apply(format_excel_dpa, axis=1)

                            def warna_baris_dpa(row):
                                lvl = df_excel_dpa.loc[row.name, "Level"]
                                if lvl == 1:
                                    return ["background-color: #ddebf7; font-weight: bold"] * len(row)
                                if lvl == 2:
                                    return ["background-color: #fff2cc; font-weight: bold"] * len(row)
                                if lvl == 3:
                                    return ["background-color: #fce4d6; font-weight: bold"] * len(row)
                                if lvl == 4:
                                    return ["background-color: #e2efda; font-weight: bold"] * len(row)
                                return [""] * len(row)

                            output_dpa = io.BytesIO()
                            with pd.ExcelWriter(output_dpa, engine="openpyxl") as writer:
                                df_excel_dpa.drop(columns=["Level"]).style.apply(warna_baris_dpa, axis=1).to_excel(
                                    writer, index=False, sheet_name="Integrasi_DPA"
                                )
                            output_dpa.seek(0)

                            st.download_button(
                                label="📥 Download Excel (Link DPA)",
                                data=output_dpa,
                                file_name=f"Integrasi_DPA_{nama_file_export}_{tahun_pilihan}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                key="dl_t3",
                            )
                    except Exception as e:
                        st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")

    # -------------------------------------------------------------------
    # TAB 4: EVALUASI REALISASI
    # -------------------------------------------------------------------
    with tab4:
        st.info(
            f"💡 Patokan pagu anggaran menggunakan tahapan: **{tahap_akhir}**. "
            "Anda bisa mengosongkan salah satu input jika tidak tersedia."
        )

        sumber_data = st.radio(
            "Pilih Mode Input Data",
            ["📂 Upload File Lokal (Excel/CSV)", "🌐 Link Google Sheet (Public)"],
            horizontal=True,
        )

        file_realisasi = None
        file_pptk = None
        link_realisasi = ""
        link_pptk = ""

        col_up1, col_up2 = st.columns(2)

        if sumber_data == "📂 Upload File Lokal (Excel/CSV)":
            with col_up1:
                st.markdown("**1️⃣ Data Realisasi Keuangan**")
                file_realisasi = st.file_uploader(
                    "Upload Excel (kolom wajib: 'kode sub', 'realisasi')",
                    type=["xlsx", "xls", "csv"],
                    key="up_realisasi",
                )
            with col_up2:
                st.markdown("**2️⃣ Master Bidang / PPTK**")
                file_pptk = st.file_uploader(
                    "Upload Excel (kolom wajib: 'kode sub', 'penanggung jawab')",
                    type=["xlsx", "xls", "csv"],
                    key="up_pptk",
                )
        else:
            with col_up1:
                st.markdown("**1️⃣ Data Realisasi Keuangan**")
                link_realisasi = st.text_input(
                    "🔗 Paste Link Google Sheet Realisasi",
                    placeholder="https://docs.google.com/spreadsheets/d/...",
                )
                st.caption("Pastikan akses link diatur ke: Anyone with the link")
            with col_up2:
                st.markdown("**2️⃣ Master Bidang / PPTK**")
                link_pptk = st.text_input(
                    "🔗 Paste Link Google Sheet Master Bidang",
                    placeholder="https://docs.google.com/spreadsheets/d/...",
                )
                st.caption("Pastikan akses link diatur ke: Anyone with the link")

        def konversi_link_gsheet(url):
            if pd.isna(url) or str(url).strip() == "":
                return None
            url = str(url).strip()
            if "docs.google.com/spreadsheets" in url:
                match = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
                if match:
                    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv"
            return None

        if st.button("🚀 PROSES EVALUASI REALISASI", type="primary", use_container_width=True, key="btn_tab4"):
            with st.spinner("Menyedot data dan menjahit dengan pembersih karakter..."):
                df_eval = df_proses[df_proses["tahapan"] == tahap_akhir].copy()

                if df_eval.empty:
                    st.error(f"⚠️ Tidak ada data anggaran untuk tahapan {tahap_akhir}.")
                else:
                    df_base = df_eval.groupby(["kode_sub_kegiatan", "nama_sub_kegiatan"])["pagu"].sum().reset_index()
                    df_base.rename(
                        columns={
                            "kode_sub_kegiatan": "Kode Sub",
                            "nama_sub_kegiatan": "Uraian Sub Kegiatan",
                            "pagu": "Pagu Anggaran",
                        },
                        inplace=True,
                    )

                    df_base["key_merge"] = df_base["Kode Sub"].astype(str).str.replace(r"[^0-9.]", "", regex=True)

                    df_real = pd.DataFrame()
                    try:
                        if sumber_data == "📂 Upload File Lokal (Excel/CSV)" and file_realisasi is not None:
                            df_real = pd.read_csv(file_realisasi) if file_realisasi.name.endswith(".csv") else pd.read_excel(file_realisasi)
                        elif sumber_data == "🌐 Link Google Sheet (Public)" and link_realisasi != "":
                            url_csv = konversi_link_gsheet(link_realisasi)
                            if url_csv:
                                df_real = pd.read_csv(url_csv)
                    except Exception as e:
                        st.error(f"❌ Gagal menarik data Realisasi: {e}")

                    if not df_real.empty:
                        df_real.columns = df_real.columns.astype(str).str.lower().str.strip()
                        if "kode sub" in df_real.columns and "realisasi" in df_real.columns:
                            df_real["key_merge"] = df_real["kode sub"].astype(str).str.replace(r"[^0-9.]", "", regex=True)

                            angka_bersih = (
                                df_real["realisasi"]
                                .astype(str)
                                .str.replace(r"[Rp\s\.]", "", regex=True)
                                .str.replace(",", ".")
                            )
                            df_real["Realisasi"] = pd.to_numeric(angka_bersih, errors="coerce").fillna(0)

                            df_real = df_real.groupby("key_merge")["Realisasi"].sum().reset_index()
                            df_base = pd.merge(df_base, df_real, on="key_merge", how="left")
                            df_base["Realisasi"] = df_base["Realisasi"].fillna(0)
                        else:
                            st.warning("⚠️ Kolom 'kode sub' atau 'realisasi' tidak ditemukan di data Realisasi.")
                            df_base["Realisasi"] = 0
                    else:
                        df_base["Realisasi"] = 0

                    df_pj = pd.DataFrame()
                    try:
                        if sumber_data == "📂 Upload File Lokal (Excel/CSV)" and file_pptk is not None:
                            df_pj = pd.read_csv(file_pptk) if file_pptk.name.endswith(".csv") else pd.read_excel(file_pptk)
                        elif sumber_data == "🌐 Link Google Sheet (Public)" and link_pptk != "":
                            url_csv = konversi_link_gsheet(link_pptk)
                            if url_csv:
                                df_pj = pd.read_csv(url_csv)
                    except Exception as e:
                        st.error(f"❌ Gagal menarik data Master Bidang: {e}")

                    if not df_pj.empty:
                        df_pj.columns = df_pj.columns.astype(str).str.lower().str.strip()
                        if "kode sub" in df_pj.columns and "penanggung jawab" in df_pj.columns:
                            df_pj["key_merge"] = df_pj["kode sub"].astype(str).str.replace(r"[^0-9.]", "", regex=True)
                            df_pj["Penanggung Jawab"] = (
                                df_pj["penanggung jawab"].astype(str).replace(["nan", "NaN", "None", ""], "BELUM DIPETAKAN")
                            )
                            df_pj = df_pj.drop_duplicates(subset=["key_merge"])
                            df_pj = df_pj[["key_merge", "Penanggung Jawab"]]
                            df_base = pd.merge(df_base, df_pj, on="key_merge", how="left")
                            df_base["Penanggung Jawab"] = df_base["Penanggung Jawab"].fillna("BELUM DIPETAKAN")
                        else:
                            st.warning("⚠️ Kolom 'kode sub' atau 'penanggung jawab' tidak ditemukan di Master Bidang.")
                            df_base["Penanggung Jawab"] = "BELUM DIPETAKAN"
                    else:
                        df_base["Penanggung Jawab"] = "BELUM DIPETAKAN"

                    df_base["Sisa Anggaran"] = df_base["Pagu Anggaran"] - df_base["Realisasi"]
                    df_base["% Capaian"] = (df_base["Realisasi"] / df_base["Pagu Anggaran"].replace(0, pd.NA)).fillna(0) * 100

                    df_base = df_base.sort_values(by=["Penanggung Jawab", "Kode Sub"]).reset_index(drop=True)
                    kolom_urut = [
                        "Kode Sub",
                        "Uraian Sub Kegiatan",
                        "Pagu Anggaran",
                        "Realisasi",
                        "Sisa Anggaran",
                        "% Capaian",
                        "Penanggung Jawab",
                    ]
                    df_final_eval = df_base[kolom_urut]

                    st.success("✅ Evaluasi Realisasi berhasil dibuat!")
                    st.dataframe(
                        df_final_eval,
                        use_container_width=True,
                        height=500,
                        column_config={
                            "Pagu Anggaran": st.column_config.NumberColumn(format="%.0f"),
                            "Realisasi": st.column_config.NumberColumn(format="%.0f"),
                            "Sisa Anggaran": st.column_config.NumberColumn(format="%.0f"),
                            "% Capaian": st.column_config.NumberColumn(format="%.2f %%"),
                        },
                    )

                    output_eval = io.BytesIO()
                    with pd.ExcelWriter(output_eval, engine="openpyxl") as writer:
                        format_eval = {
                            "Pagu Anggaran": "{:,.0f}",
                            "Realisasi": "{:,.0f}",
                            "Sisa Anggaran": "{:,.0f}",
                            "% Capaian": "{:.2f}",
                        }
                        df_final_eval.style.format(format_eval).to_excel(
                            writer, index=False, sheet_name="Evaluasi_Realisasi"
                        )
                    output_eval.seek(0)

                    st.download_button(
                        label="📥 Download Excel (Evaluasi Realisasi)",
                        data=output_eval,
                        file_name=f"Evaluasi_Realisasi_{nama_file_export}_{tahun_pilihan}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        key="dl_t4",
                    )

    # -------------------------------------------------------------------
    # TAB 5: REKAP PER BIDANG
    # -------------------------------------------------------------------
    with tab5:
        st.info(f"💡 Menampilkan total pagu per Bidang/PPTK: **{tahap_awal}** vs **{tahap_akhir}**")

        sumber_data_bidang = st.radio(
            "Pilih Mode Input File Pemetaan PPTK/Bidang",
            ["📂 Upload File Lokal (Excel/CSV)", "🌐 Link Google Sheet (Otomatis Baca Sheet)"],
            horizontal=True,
            key="radio_bidang_t5",
        )

        file_mapping_bidang = None
        link_bidang_input = ""
        df_map_gsheet = pd.DataFrame()

        if sumber_data_bidang == "📂 Upload File Lokal (Excel/CSV)":
            file_mapping_bidang = st.file_uploader(
                "📂 Upload File Excel Pemetaan (harus ada kolom 'kode sub' dan 'penanggung jawab')",
                type=["xlsx", "xls", "csv"],
                key="up_bidang_t5",
            )
        else:
            link_bidang_input = st.text_input(
                "🔗 Paste Link Google Sheet Pemetaan",
                placeholder="https://docs.google.com/spreadsheets/d/...",
                key="link_bidang_t5",
            )
            st.caption("Gunakan link share biasa. Pastikan akses diatur ke: Anyone with the link")

            if link_bidang_input:
                match = re.search(r"/d/([a-zA-Z0-9-_]+)", link_bidang_input)
                if match:
                    doc_id = match.group(1)
                    url_xlsx = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"

                    try:
                        @st.cache_data(show_spinner=False, ttl=600)
                        def tarik_excel_bidang(url):
                            resp = requests.get(url)
                            resp.raise_for_status()
                            return resp.content

                        with st.spinner("🔍 Sedang membaca Google Sheet untuk mencari daftar sheet..."):
                            excel_bytes = tarik_excel_bidang(url_xlsx)
                            xls = pd.ExcelFile(io.BytesIO(excel_bytes))
                            daftar_sheet = xls.sheet_names

                        if daftar_sheet:
                            sheet_pilihan = st.selectbox(
                                "📑 Pilih sheet yang berisi data PPTK/Bidang",
                                daftar_sheet,
                                key="sheet_bidang_t5",
                            )
                            if sheet_pilihan:
                                df_map_gsheet = pd.read_excel(xls, sheet_name=sheet_pilihan)
                        else:
                            st.error("❌ Tidak ada sheet yang ditemukan di file tersebut.")
                    except Exception as e:
                        st.error(f"❌ Gagal membaca Google Sheet. Error: {e}")
                else:
                    st.warning("⚠️ Link tidak valid. Coba paste ulang link Google Sheet yang benar.")

        if st.button("📊 PROSES REKAP BIDANG", type="primary", use_container_width=True, key="btn_tab5"):
            if sumber_data_bidang == "📂 Upload File Lokal (Excel/CSV)" and file_mapping_bidang is None:
                st.error("⚠️ Mohon upload file Excel/CSV terlebih dahulu!")
            elif sumber_data_bidang == "🌐 Link Google Sheet (Otomatis Baca Sheet)" and link_bidang_input == "":
                st.error("⚠️ Mohon paste Link Google Sheet terlebih dahulu!")
            elif sumber_data_bidang == "🌐 Link Google Sheet (Otomatis Baca Sheet)" and df_map_gsheet.empty:
                st.error("⚠️ Menunggu data dari Google Sheet. Silakan pilih sheet yang benar.")
            else:
                with st.spinner("Menyatukan data SIPD dengan pemetaan PPTK/Bidang..."):
                    try:
                        if sumber_data_bidang == "📂 Upload File Lokal (Excel/CSV)":
                            if file_mapping_bidang.name.endswith(".csv"):
                                df_map = pd.read_csv(file_mapping_bidang)
                            else:
                                df_map = pd.read_excel(file_mapping_bidang)
                        else:
                            df_map = df_map_gsheet.copy()

                        df_map.columns = df_map.columns.astype(str).str.lower().str.strip()

                        if "code" in df_map.columns:
                            df_map.rename(columns={"code": "kode sub"}, inplace=True)
                        if "bidang" in df_map.columns:
                            df_map.rename(columns={"bidang": "penanggung jawab"}, inplace=True)

                        if "kode sub" not in df_map.columns or "penanggung jawab" not in df_map.columns:
                            st.error(
                                f"❌ File pemetaan harus memiliki kolom 'kode sub' dan 'penanggung jawab'. "
                                f"Kolom yang terdeteksi: {list(df_map.columns)}"
                            )
                        else:
                            df_map = df_map[["kode sub", "penanggung jawab"]].rename(columns={"kode sub": "kode_sub_kegiatan"})
                            df_map["kode_sub_kegiatan"] = df_map["kode_sub_kegiatan"].astype(str).str.strip()
                            df_map["penanggung jawab"] = df_map["penanggung jawab"].fillna("TIDAK ADA DATA")
                            df_map = df_map.drop_duplicates(subset=["kode_sub_kegiatan"])

                            df_sipd_filter = df_proses[df_proses["tahapan"].isin([tahap_awal, tahap_akhir])].copy()
                            df_gabung = pd.merge(df_sipd_filter, df_map, on="kode_sub_kegiatan", how="left")
                            df_gabung["penanggung jawab"] = df_gabung["penanggung jawab"].fillna("TIDAK DIPETAKAN")

                            rekap_bidang = df_gabung.groupby(["penanggung jawab", "tahapan"])["pagu"].sum().reset_index()
                            pivot_bidang = (
                                rekap_bidang.pivot_table(
                                    index="penanggung jawab",
                                    columns="tahapan",
                                    values="pagu",
                                    aggfunc="sum",
                                    fill_value=0,
                                )
                                .reset_index()
                            )

                            for t in [tahap_awal, tahap_akhir]:
                                if t not in pivot_bidang.columns:
                                    pivot_bidang[t] = 0

                            pivot_bidang["Selisih"] = pivot_bidang[tahap_akhir] - pivot_bidang[tahap_awal]
                            pivot_bidang.rename(
                                columns={
                                    "penanggung jawab": "Penanggung Jawab / Bidang",
                                    tahap_awal: f"Pagu {tahap_awal}",
                                    tahap_akhir: f"Pagu {tahap_akhir}",
                                },
                                inplace=True,
                            )

                            baris_total = pd.DataFrame(
                                [
                                    {
                                        "Penanggung Jawab / Bidang": "TOTAL KESELURUHAN",
                                        f"Pagu {tahap_awal}": pivot_bidang[f"Pagu {tahap_awal}"].sum(),
                                        f"Pagu {tahap_akhir}": pivot_bidang[f"Pagu {tahap_akhir}"].sum(),
                                        "Selisih": pivot_bidang["Selisih"].sum(),
                                    }
                                ]
                            )
                            pivot_bidang = pd.concat([pivot_bidang, baris_total], ignore_index=True)

                            st.success("✅ Rekapitulasi per bidang selesai!")
                            st.dataframe(
                                pivot_bidang,
                                use_container_width=True,
                                column_config={
                                    f"Pagu {tahap_awal}": st.column_config.NumberColumn(format="Rp %.0f"),
                                    f"Pagu {tahap_akhir}": st.column_config.NumberColumn(format="Rp %.0f"),
                                    "Selisih": st.column_config.NumberColumn(format="Rp %.0f"),
                                },
                            )

                            output_bidang = io.BytesIO()
                            with pd.ExcelWriter(output_bidang, engine="openpyxl") as writer:
                                pivot_bidang.to_excel(writer, index=False, sheet_name="Rekap_Bidang_Internal")
                            output_bidang.seek(0)

                            st.download_button(
                                label="📥 Download Excel (Rekap Bidang)",
                                data=output_bidang,
                                file_name=f"Rekap_Bidang_Internal_{nama_file_export}_{tahun_pilihan}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                key="dl_tab5_excel",
                            )
                    except Exception as e:
                        st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")

    # -------------------------------------------------------------------
    # TAB 6: REKAP KODE REKENING
    # -------------------------------------------------------------------
    with tab6:
        st.markdown("### 📦 Rekap Kode Rekening (dulu NPD)")

        list_skpd_npd = sorted([s for s in df_proses["nama_skpd"].unique() if s != ""])
        list_skpd_npd.insert(0, "SEMUA SKPD")
        skpd_npd = st.multiselect("🏢 Pilih SKPD", list_skpd_npd, default=["SEMUA SKPD"])

        if "SEMUA SKPD" in skpd_npd:
            df_npd = df_proses.copy()
        else:
            df_npd = df_proses[df_proses["nama_skpd"].isin(skpd_npd)].copy()

        list_sub_npd = sorted([s for s in df_npd["nama_sub_kegiatan"].unique() if s != ""])
        sub_npd = st.multiselect("🔍 Pilih Sub Kegiatan", list_sub_npd, default=list_sub_npd)

        if sub_npd:
            df_npd = df_npd[df_npd["nama_sub_kegiatan"].isin(sub_npd)]

        if df_npd.empty:
            st.warning("Tidak ada data untuk pilihan filter ini.")
        else:
            df_npd_rek = df_npd[df_npd["kode_rekening"] != ""].copy()
            df_npd_rek["Major Rek"] = df_npd_rek["kode_rekening"].str.slice(0, 5)

            rekap_npd = (
                df_npd_rek.groupby(
                    ["Major Rek", "nama_rekening", "nama_skpd", "nama_sub_kegiatan", "tahapan"]
                )["pagu"]
                .sum()
                .reset_index()
            )

            pivot_npd = (
                rekap_npd.pivot_table(
                    index=["Major Rek", "nama_rekening", "nama_skpd", "nama_sub_kegiatan"],
                    columns="tahapan",
                    values="pagu",
                    aggfunc="sum",
                    fill_value=0,
                )
                .reset_index()
            )

            for t in list_tahapan:
                if t not in pivot_npd.columns:
                    pivot_npd[t] = 0

            pivot_npd["Selisih"] = pivot_npd[tahap_akhir] - pivot_npd[tahap_awal]

            urut_npd = ["Major Rek", "nama_rekening", "nama_skpd", "nama_sub_kegiatan"] + list_tahapan + ["Selisih"]
            for col in urut_npd:
                if col not in pivot_npd.columns:
                    pivot_npd[col] = 0
            pivot_npd = pivot_npd[urut_npd]

            st.dataframe(
                pivot_npd,
                use_container_width=True,
                column_config={t: st.column_config.NumberColumn(format="Rp %.0f") for t in list_tahapan},
            )

            output_npd = io.BytesIO()
            with pd.ExcelWriter(output_npd, engine="openpyxl") as writer:
                pivot_npd.to_excel(writer, index=False, sheet_name="Rekap_Kode_Rekening")
            output_npd.seek(0)

            st.download_button(
                label="📥 Download Rekap Kode Rekening (Excel)",
                data=output_npd,
                file_name=f"Rekap_Kode_Rekening_{tahun_pilihan}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
